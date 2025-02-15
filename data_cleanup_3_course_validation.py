import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 1. Load the check-in form
check_in_form = pd.read_excel('name_id_mismatch_second_check.xlsx')

# 2. Load the combined roster (previously created by merging all rosters)
combined_roster = pd.read_excel('combined_Roster.xlsx')

# 3. Standardize column names
check_in_form.columns = check_in_form.columns.str.lower().str.strip()
combined_roster.columns = combined_roster.columns.str.lower().str.strip()

# 4. Ensure IDs are strings and clean them
check_in_form['student id'] = (
    check_in_form['student id']
    .astype(str)
    .str.strip()
    .str.split('.')
    .str[0]  # Removes any ".0" suffix
)

combined_roster['id'] = combined_roster['id'].astype(str).str.strip()

# 5. Split "what course do you need assistance with?" into 'subject' and 'catalog'
check_in_form[['subject', 'catalog']] = check_in_form['what course do you need assistance with?'].str.split(' ', n=1, expand=True)

# 6. Convert subject/catalog to uppercase and strip spaces
check_in_form['subject'] = check_in_form['subject'].str.upper().str.strip()
check_in_form['catalog'] = check_in_form['catalog'].astype(str).str.strip()

combined_roster['subject'] = combined_roster['subject'].str.upper().str.strip()
combined_roster['catalog'] = combined_roster['catalog'].astype(str).str.strip()

# 7. Filter rows where students are here to see a tutor
tutoring_sessions = check_in_form[check_in_form['i am here to:'] == 'See a tutor']

# 8. Prepare to highlight mismatched rows in Excel
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Load the Excel file we want to modify
wb = load_workbook('name_id_mismatch_second_check.xlsx')
ws = wb.active

# 9. Verify each tutoring session
for idx, row in tutoring_sessions.iterrows():
    student_id = row['student id']
    subject = row['subject']
    catalog = row['catalog']

    # Filter the combined roster for matching subject/catalog
    matching_roster = combined_roster[
        (combined_roster['subject'] == subject) & 
        (combined_roster['catalog'] == catalog)
    ]

    # Check if this student is found in that subset
    student_match = matching_roster[matching_roster['id'] == student_id]

    if student_match.empty:
        # If empty, the student is not enrolled in that class
        for cell in ws[idx + 2]:  # +2 accounts for 0-based index + header row
            cell.fill = yellow_fill

# 10. Save the updated Excel file
wb.save('course_validation.xlsx')
print("Validation complete. Rows for students not enrolled in their tutoring class have been highlighted.")
