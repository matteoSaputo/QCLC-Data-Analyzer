from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
from IPython.display import display

# Load the data
check_in_form = pd.read_excel('name_id_mismatch.xlsx')
all_enrolled = pd.read_excel('all_enrolled.xlsx')

# Standardize column names and data
check_in_form.columns = check_in_form.columns.str.lower().str.strip()
all_enrolled.columns = all_enrolled.columns.str.lower().str.strip()

# Ensure names and IDs are strings
check_in_form['student id'] = check_in_form['student id'].astype(str).str.strip()
check_in_form['first name'] = check_in_form['first name'].str.strip().str.lower()
check_in_form['last name'] = check_in_form['last name'].str.strip().str.lower()

all_enrolled['student id'] = all_enrolled['student id'].astype(str).str.strip()
all_enrolled['first name'] = all_enrolled['first name'].str.strip().str.lower()
all_enrolled['last name'] = all_enrolled['last name'].str.strip().str.lower()

# Clean student IDs by removing '.0' if it exists
check_in_form['student id'] = check_in_form['student id'].str.split('.').str[0]
all_enrolled['student id'] = all_enrolled['student id'].str.split('.').str[0]


display(check_in_form)

display(all_enrolled)

# Merge on 'student id' and other unique identifiers
merged = check_in_form.merge(all_enrolled, 
                             on=['student id'], 
                             how='left', 
                             indicator=True)

# Identify rows not matched in 'all enrolled'
unmatched = merged[merged['_merge'] == 'left_only']


display(merged)

display(unmatched)

# Add a flag for unmatched rows
check_in_form['mismatch'] = check_in_form['student id'].isin(unmatched['student id'])


display(check_in_form)

# Save the DataFrame to an Excel file
output_file = 'name_id_mismatch_second_check.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    check_in_form.to_excel(writer, index=False, sheet_name='Check In Form')

# Load the workbook for formatting
workbook = load_workbook(output_file)
worksheet = workbook['Check In Form']

# Define red fill for formatting
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply red fill to rows with mismatches
for row_idx, mismatch in enumerate(check_in_form['mismatch'], start=2):  # Start=2 to account for header row
    if mismatch:  # If mismatch is True
        for col_idx in range(1, len(check_in_form.columns) + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
    

# Save the workbook
workbook.save(output_file)
workbook.close()

# Load the first output (name_id_mismatch.xlsx) to extract formatting
wb_first_run = load_workbook('name_id_mismatch.xlsx')
ws_first_run = wb_first_run.active

# Load the second output (name_id_mismatch_second_check.xlsx) where we will apply the formatting
wb_second_run = load_workbook('name_id_mismatch_second_check.xlsx')
ws_second_run = wb_second_run.active

# Define yellow fill color (ensure this matches the color you're checking for)
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Iterate through the rows of the first run (to find yellow-colored rows)
yellow_rows = []
for row_idx, row in enumerate(ws_first_run.iter_rows(min_row=2, max_row=ws_first_run.max_row), start=2):  # Start from row 2 to skip header
    # Check if the first cell in the row is yellow
    if row[0].fill.start_color.rgb == "FFFF00":  # Check if the first cell is yellow
        yellow_rows.append(row_idx)

# Apply the yellow color to the same rows in the second run
for row_idx in yellow_rows:
    row = ws_second_run[row_idx]
    for cell in row:
        cell.fill = yellow_fill

# Save the updated second run file with the preserved formatting
wb_second_run.save('name_id_mismatch_second_check.xlsx')