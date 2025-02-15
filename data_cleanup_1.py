import pandas as pd
from IPython.display import display
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the data
check_in_form = pd.read_excel('check_in_form.xlsx')
all_enrolled = pd.read_excel('all_enrolled.xlsx')


display(check_in_form)

display(all_enrolled)

# Standardize columns (lowercase and strip whitespaces)
check_in_form.columns = check_in_form.columns.str.lower().str.strip()
all_enrolled.columns = all_enrolled.columns.str.lower().str.strip()

# Ensure names and IDs are strings
check_in_form['student id'] = check_in_form['student id'].astype(str).str.strip()
check_in_form['first name'] = check_in_form['first name'].str.strip().str.lower()
check_in_form['last name'] = check_in_form['last name'].str.strip().str.lower()

all_enrolled['student id'] = all_enrolled['student id'].astype(str).str.strip()
all_enrolled['first name'] = all_enrolled['first name'].str.strip().str.lower()
all_enrolled['last name'] = all_enrolled['last name'].str.strip().str.lower()

# Clean student IDs by removing '.0' if it exists
check_in_form['student id'] = check_in_form['student id'].str.split('.').str[0]
all_enrolled['student id'] = all_enrolled['student id'].str.split('.').str[0]

# Merge on 'student id' and other unique identifiers
merged = check_in_form.merge(all_enrolled, 
                             on=['student id', 'first name', 'last name'], 
                             how='left', 
                             indicator=True)

# Identify rows not matched in 'all enrolled'
unmatched = merged[merged['_merge'] == 'left_only']


display(unmatched)

# Add a flag for unmatched rows
check_in_form['mismatch'] = check_in_form['student id'].isin(unmatched['student id'])

# Save the DataFrame to an Excel file
output_file = 'name_id_mismatch.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    check_in_form.to_excel(writer, index=False, sheet_name='Check In Form')

# Load the workbook for formatting
workbook = load_workbook(output_file)
worksheet = workbook['Check In Form']

# Define red fill for formatting
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply red fill to rows with mismatches
for row_idx, mismatch in enumerate(check_in_form['mismatch'], start=2):  # Start=2 to account for header row
    if mismatch:  # If mismatch is True
        for col_idx in range(1, len(check_in_form.columns) + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

# Save the workbook
workbook.save(output_file)
workbook.close()